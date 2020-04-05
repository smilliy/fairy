# -*- coding: utf-8 -*-

# -*- coding: utf-8 -*-

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


def cell_format(ws, row=0, column=0):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    alignment = Alignment(horizontal="left", vertical="center")


    first_cell = ws.cell(column=column, row=row)
    first_cell.alignment = alignment

    first_cell.border = top + left + right + bottom


if __name__ == '__main__':
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    my_cell = ws['B2']
    my_cell.value = "My Cell"

    double = Side(border_style="double", color="ff0000")


    fill = PatternFill("solid", fgColor="DDDDDD")
    fill = GradientFill(stop=("000000", "FFFFFF"))
    font = Font(b=False, color="000000")
    al = Alignment(horizontal="center", vertical="center")

    cell_format(ws, row=2, column=2)
    wb.save("styled.xlsx")